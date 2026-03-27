"""Tests for the TEA simulation and XLSX generation.

The simulation test runs the full BioSTEAM corn stover model and validates
against published NREL benchmarks.
"""

import warnings

import pytest

from processflow.schema.process_spec import ProcessSpec


class TestSimulation:
    @pytest.fixture(autouse=True)
    def suppress_warnings(self):
        warnings.filterwarnings("ignore")
        yield

    def test_corn_stover_simulation_runs(self, corn_stover_spec: ProcessSpec):
        """Corn stover simulation completes without error."""
        from processflow.tea.simulation import run_cornstover_simulation

        results = run_cornstover_simulation(corn_stover_spec)
        assert results.mesp_usd_per_kg > 0
        assert results.tci_usd > 0
        assert len(results.unit_costs) > 0

    def test_mesp_benchmark(self, corn_stover_spec: ProcessSpec):
        """MESP is within 15% of the NREL published benchmark ($2.15/gal)."""
        from processflow.tea.simulation import run_cornstover_simulation

        results = run_cornstover_simulation(corn_stover_spec)
        nrel_benchmark = 2.15  # $/gal
        tolerance = 0.15  # 15%
        lower = nrel_benchmark * (1 - tolerance)
        upper = nrel_benchmark * (1 + tolerance)
        assert lower <= results.mesp_usd_per_gal <= upper, (
            f"MESP ${results.mesp_usd_per_gal:.2f}/gal is outside "
            f"15% of NREL benchmark ${nrel_benchmark}/gal "
            f"(expected ${lower:.2f} - ${upper:.2f})"
        )

    def test_unsupported_process_raises(self):
        """Non-corn-stover processes raise NotImplementedError."""
        from processflow.schema.process_spec import Feedstock, Stream, UnitOperation, UnitType
        from processflow.tea.simulation import run_simulation

        spec = ProcessSpec(
            process_name="Biodiesel from Soybean Oil",
            feedstock=Feedstock(name="soybean_oil", flow_rate_kg_hr=5000),
            units=[UnitOperation(id="U-101", type=UnitType.REACTOR)],
            streams=[Stream(from_id="feed", to_id="U-101")],
        )
        with pytest.raises(NotImplementedError):
            run_simulation(spec)

    def test_cashflow_table_present(self, corn_stover_spec: ProcessSpec):
        """Simulation results include a cashflow table."""
        from processflow.tea.simulation import run_cornstover_simulation

        results = run_cornstover_simulation(corn_stover_spec)
        assert results.cashflow_table is not None
        assert results.cashflow_table.shape[0] > 0
        assert results.cashflow_table.shape[1] > 0


class TestXLSXWriter:
    @pytest.fixture(autouse=True)
    def suppress_warnings(self):
        warnings.filterwarnings("ignore")
        yield

    def test_xlsx_has_all_sheets(self, corn_stover_spec: ProcessSpec, tmp_output):
        """Generated XLSX contains all 8 required sheets."""
        from openpyxl import load_workbook

        from processflow.tea.simulation import run_cornstover_simulation
        from processflow.tea.xlsx_writer import write_tea_xlsx

        results = run_cornstover_simulation(corn_stover_spec)
        path = write_tea_xlsx(results, corn_stover_spec, tmp_output / "test.xlsx")

        wb = load_workbook(str(path))
        expected_sheets = [
            "Summary", "Process Inputs", "Mass Balance", "Energy Balance",
            "Equipment Costs", "Operating Costs", "Cash Flow", "Sensitivity",
        ]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

    def test_summary_sheet_has_mesp(self, corn_stover_spec: ProcessSpec, tmp_output):
        """Summary sheet contains MESP value."""
        from openpyxl import load_workbook

        from processflow.tea.simulation import run_cornstover_simulation
        from processflow.tea.xlsx_writer import write_tea_xlsx

        results = run_cornstover_simulation(corn_stover_spec)
        path = write_tea_xlsx(results, corn_stover_spec, tmp_output / "test.xlsx")

        wb = load_workbook(str(path))
        ws = wb["Summary"]
        # Find MESP row
        found_mesp = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value and "MESP" in str(cell.value):
                    found_mesp = True
                    break
        assert found_mesp, "MESP not found in Summary sheet"
