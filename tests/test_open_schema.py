"""Tests for open schema (arbitrary unit types) and generic TEA calculator."""

import json

import pytest

from processflow.schema.process_spec import (
    ChemicalRole,
    ChemicalSpec,
    EconomicParams,
    Feedstock,
    ProcessSpec,
    Product,
    Stream,
    UnitOperation,
    UnitType,
)


class TestOpenUnitType:
    def test_arbitrary_string_type_accepted(self):
        """Custom unit type strings are accepted without error."""
        spec = ProcessSpec(
            process_name="CCS Process",
            feedstock=Feedstock(name="exhaust_gas", flow_rate_kg_hr=45000),
            units=[UnitOperation(id="U-101", type="sox_prescrubber")],
            streams=[Stream(from_id="feed", to_id="U-101")],
        )
        assert spec.units[0].type == "sox_prescrubber"

    def test_enum_type_coerced_to_string(self):
        """UnitType enum values are coerced to their string representation."""
        unit = UnitOperation(id="U-101", type=UnitType.MIXER)
        assert unit.type == "Mixer"
        assert isinstance(unit.type, str)

    def test_multiple_custom_types(self):
        """Spec with multiple custom unit types loads correctly."""
        spec = ProcessSpec(
            process_name="Maritime CCS",
            feedstock=Feedstock(name="exhaust", flow_rate_kg_hr=45000),
            units=[
                UnitOperation(id="U-001", type="sox_prescrubber", name="SOx Scrubber"),
                UnitOperation(id="U-002", type="amine_absorber", name="MEA Absorber"),
                UnitOperation(id="U-003", type="co2_compressor", name="CO2 Compressor"),
            ],
            streams=[
                Stream(from_id="FEED", to_id="U-001"),
                Stream(from_id="U-001", to_id="U-002"),
                Stream(from_id="U-002", to_id="U-003"),
            ],
        )
        assert len(spec.units) == 3
        assert spec.units[1].type == "amine_absorber"


class TestPreprocessor:
    def test_price_per_ton_normalized(self):
        """price_per_ton is mapped to price_usd_per_ton."""
        data = {
            "process_name": "Test",
            "feedstock": {"name": "x", "flow_rate_kg_hr": 100, "price_per_ton": 50.0},
            "units": [{"id": "U-1", "type": "Mixer"}],
            "streams": [{"from_id": "feed", "to_id": "U-1"}],
        }
        spec = ProcessSpec.from_json(json.dumps(data))
        assert spec.feedstock.price_usd_per_ton == 50.0

    def test_yield_kg_hr_normalized(self):
        """yield_kg_hr is mapped to expected_yield_kg_hr."""
        data = {
            "process_name": "Test",
            "feedstock": {"name": "x", "flow_rate_kg_hr": 100},
            "products": [{"name": "CO2", "yield_kg_hr": 99.0}],
            "units": [{"id": "U-1", "type": "Mixer"}],
            "streams": [{"from_id": "feed", "to_id": "U-1"}],
        }
        spec = ProcessSpec.from_json(json.dumps(data))
        assert spec.products[0].expected_yield_kg_hr == 99.0

    def test_parameters_normalized_to_params(self):
        """parameters field is mapped to params."""
        data = {
            "process_name": "Test",
            "feedstock": {"name": "x", "flow_rate_kg_hr": 100},
            "units": [{"id": "U-1", "type": "Mixer", "parameters": {"T": 300}}],
            "streams": [{"from_id": "feed", "to_id": "U-1"}],
        }
        spec = ProcessSpec.from_json(json.dumps(data))
        assert spec.units[0].params == {"T": 300}

    def test_tax_rate_normalized(self):
        """tax_rate is mapped to income_tax_rate."""
        data = {
            "process_name": "Test",
            "feedstock": {"name": "x", "flow_rate_kg_hr": 100},
            "units": [{"id": "U-1", "type": "Mixer"}],
            "streams": [{"from_id": "feed", "to_id": "U-1"}],
            "economic": {"tax_rate": 0.0, "operating_days": 300},
        }
        spec = ProcessSpec.from_json(json.dumps(data))
        assert spec.economic.income_tax_rate == 0.0


class TestExtraFields:
    def test_chemical_extra_fields_preserved(self):
        """Extra fields on ChemicalSpec are preserved via extra='allow'."""
        chem = ChemicalSpec(
            name="MEA",
            role=ChemicalRole.CATALYST,
            price_per_kg=1.80,
            consumption_rate_kg_hr=0.15,
        )
        assert chem.model_extra["price_per_kg"] == 1.80
        assert chem.model_extra["consumption_rate_kg_hr"] == 0.15

    def test_economic_extra_fields_preserved(self):
        """Extra fields on EconomicParams are preserved."""
        eco = EconomicParams(
            operating_days=300,
            capex_usd=8_000_000,
            annual_maintenance_usd=150_000,
        )
        assert eco.capex_usd == 8_000_000
        assert eco.model_extra["annual_maintenance_usd"] == 150_000

    def test_feedstock_notes_preserved(self):
        """Notes on Feedstock are preserved as extra."""
        fs = Feedstock(
            name="Exhaust Gas",
            flow_rate_kg_hr=45000,
            notes="From 10 MW marine diesel",
        )
        assert fs.model_extra["notes"] == "From 10 MW marine diesel"


class TestStreamComponents:
    def test_dict_components_accepted(self):
        """Stream components as dict[str, float] are accepted."""
        stream = Stream(
            from_id="U-1",
            to_id="U-2",
            components={"CO2": 0.95, "H2O": 0.05},
        )
        assert isinstance(stream.components, dict)
        assert stream.component_names == ["CO2", "H2O"]

    def test_list_components_still_work(self):
        """Stream components as list[str] still work."""
        stream = Stream(
            from_id="U-1",
            to_id="U-2",
            components=["ethanol", "water"],
        )
        assert stream.component_names == ["ethanol", "water"]

    def test_none_components(self):
        """Stream with no components returns empty list from component_names."""
        stream = Stream(from_id="U-1", to_id="U-2")
        assert stream.component_names == []


class TestBoundaryNodes:
    def test_uppercase_feed_accepted(self):
        """FEED (uppercase) is accepted as a boundary node."""
        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="FEED", to_id="U-1")],
        )
        assert spec.streams[0].from_id == "FEED"

    def test_custom_boundary_nodes(self):
        """Custom boundary nodes like ENGINE and VENT are accepted."""
        spec = ProcessSpec(
            process_name="CCS",
            feedstock=Feedstock(name="exhaust", flow_rate_kg_hr=50000),
            units=[UnitOperation(id="U-1", type="amine_absorber")],
            streams=[
                Stream(from_id="ENGINE", to_id="U-1"),
                Stream(from_id="U-1", to_id="VENT"),
            ],
        )
        assert len(spec.streams) == 2


class TestGenericTEA:
    def test_basic_npv_calculation(self):
        """NPV is computed correctly from CAPEX and annual cash flows."""
        from processflow.tea.generic_tea import run_generic_tea

        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=1_000_000,
                annual_costs={"opex": 100_000},
                annual_revenues={"sales": 250_000},
                discount_rate=0.10,
                plant_lifetime_years=10,
            ),
        )
        results = run_generic_tea(spec)

        assert results.capex_usd == 1_000_000
        assert results.total_annual_costs_usd == 100_000
        assert results.total_annual_revenues_usd == 250_000
        assert results.net_annual_cashflow_usd == 150_000

        # NPV should be -CAPEX + PV of 150k/yr for 10yr at 10%
        # PV annuity = 150000 * (1 - 1.1^-10) / 0.1 ≈ 921,685
        # NPV ≈ -78,315
        assert results.npv_usd == pytest.approx(-78_349, rel=0.01)

    def test_payback_calculation(self):
        """Simple payback = CAPEX / net annual operating income."""
        from processflow.tea.generic_tea import run_generic_tea

        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=500_000,
                annual_costs={"opex": 50_000},
                annual_revenues={"sales": 150_000},
                discount_rate=0.08,
                plant_lifetime_years=15,
            ),
        )
        results = run_generic_tea(spec)
        assert results.simple_payback_years == pytest.approx(5.0, rel=0.01)

    def test_negative_net_has_no_payback(self):
        """When costs exceed revenues, payback is None."""
        from processflow.tea.generic_tea import run_generic_tea

        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=8_000_000,
                annual_costs={"fuel": 445_000, "maintenance": 150_000},
                annual_revenues={"ets_savings": 167_000},
                discount_rate=0.08,
                plant_lifetime_years=15,
            ),
        )
        results = run_generic_tea(spec)
        assert results.simple_payback_years is None
        assert results.npv_usd < 0  # deeply negative

    def test_cashflow_table_shape(self):
        """Cash flow table has one row per year of plant lifetime."""
        from processflow.tea.generic_tea import run_generic_tea

        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=1_000_000,
                annual_costs={"opex": 100_000},
                annual_revenues={"sales": 200_000},
                plant_lifetime_years=20,
            ),
        )
        results = run_generic_tea(spec)
        assert results.cashflow_table is not None
        assert results.cashflow_table.shape[0] == 20

    def test_zero_discount_rate(self):
        """CRF handles zero discount rate without division by zero."""
        from processflow.tea.generic_tea import capital_recovery_factor

        crf = capital_recovery_factor(0.0, 20)
        assert crf == pytest.approx(0.05)  # 1/20

    def test_lcop_from_product_yield(self):
        """LCOP is computed from first product's expected yield."""
        from processflow.tea.generic_tea import run_generic_tea

        spec = ProcessSpec(
            process_name="Test",
            feedstock=Feedstock(name="x", flow_rate_kg_hr=100),
            products=[Product(name="CO2", expected_yield_kg_hr=99.0)],
            units=[UnitOperation(id="U-1", type="Mixer")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=8_000_000,
                annual_costs={"opex": 891_000},
                annual_revenues={},
                discount_rate=0.08,
                plant_lifetime_years=15,
                operating_days=300,
            ),
        )
        results = run_generic_tea(spec)
        assert results.lcop_usd_per_unit is not None
        assert results.lcop_unit == "$/tonne"
        # Annual output = 99 kg/hr * 300 days * 24 hr / 1000 = 712.8 tonnes/yr
        assert results.lcop_usd_per_unit > 0


class TestGenericTEAXLSX:
    def test_generic_xlsx_has_sheets(self, tmp_path):
        """Generic TEA XLSX contains expected sheets."""
        from openpyxl import load_workbook

        from processflow.tea.generic_tea import run_generic_tea
        from processflow.tea.xlsx_writer import write_generic_tea_xlsx

        spec = ProcessSpec(
            process_name="Maritime CCS Test",
            feedstock=Feedstock(name="exhaust", flow_rate_kg_hr=45000),
            units=[UnitOperation(id="U-1", type="sox_prescrubber")],
            streams=[Stream(from_id="feed", to_id="U-1")],
            economic=EconomicParams(
                capex_usd=8_000_000,
                annual_costs={"maintenance": 150_000, "solvent": 95_000},
                annual_revenues={"ets_savings": 167_000},
                operating_days=300,
                plant_lifetime_years=15,
                discount_rate=0.08,
            ),
        )
        results = run_generic_tea(spec)
        path = write_generic_tea_xlsx(results, spec, tmp_path / "tea.xlsx")

        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames
        assert "OPEX Breakdown" in wb.sheetnames
        assert "Cash Flow" in wb.sheetnames
