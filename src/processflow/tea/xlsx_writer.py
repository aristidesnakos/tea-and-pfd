"""TEA XLSX Report Generator — creates formatted spreadsheet from simulation results.

Generates an 8-sheet Excel workbook:
1. Summary - Key metrics dashboard
2. Process Inputs - All parameters with source annotations
3. Mass Balance - Stream table
4. Energy Balance - Utility requirements
5. Equipment Costs - Purchase + installed per unit
6. Operating Costs - Annual breakdown
7. Cash Flow - Year-by-year NPV analysis
8. Sensitivity - Parameter impact data (placeholder for MVP)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from processflow.schema.process_spec import ProcessSpec
from processflow.tea.simulation import SimulationResults

# Style constants
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CURRENCY_FORMAT = '#,##0'
CURRENCY_M_FORMAT = '#,##0.0'
PERCENT_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0.0'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
SOURCE_FONT = Font(italic=True, size=9, color="888888")


def _write_header(ws: Any, row: int, headers: list[str]) -> None:
    """Write a formatted header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_subheader(ws: Any, row: int, text: str, cols: int = 5) -> None:
    """Write a section subheader spanning multiple columns."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    for col in range(2, cols + 1):
        ws.cell(row=row, column=col).fill = SUBHEADER_FILL


def _safe_value(value: Any) -> Any:
    """Convert non-scalar values to strings for Excel compatibility."""
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _write_metric_row(ws: Any, row: int, label: str, value: Any,
                       unit: str = "", fmt: str | None = None,
                       source: str = "") -> None:
    """Write a label-value-unit-source row."""
    ws.cell(row=row, column=1, value=label)
    cell = ws.cell(row=row, column=2, value=_safe_value(value))
    if fmt:
        cell.number_format = fmt
    ws.cell(row=row, column=3, value=unit)
    if source:
        src_cell = ws.cell(row=row, column=4, value=source)
        src_cell.font = SOURCE_FONT
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _auto_width(ws: Any) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_summary_sheet(wb: Workbook, results: SimulationResults, spec: ProcessSpec) -> None:
    """Sheet 1: Summary — Key metrics dashboard."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.cell(row=1, column=1, value=f"TEA Summary: {spec.process_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=spec.description).font = Font(size=9, color="666666")

    _write_header(ws, 4, ["Metric", "Value", "Unit", "Source"])

    row = 5
    _write_subheader(ws, row, "Key Performance Indicators", 4)
    row += 1

    metrics = [
        ("Minimum Ethanol Selling Price (MESP)", results.mesp_usd_per_gal, "$/gal", CURRENCY_M_FORMAT, "BioSTEAM"),
        ("MESP", results.mesp_usd_per_kg, "$/kg", CURRENCY_M_FORMAT, "BioSTEAM"),
        ("Internal Rate of Return (IRR)", results.irr, "", PERCENT_FORMAT, "BioSTEAM"),
        ("Net Present Value (NPV)", results.npv_usd / 1e6, "MM$", CURRENCY_M_FORMAT, "BioSTEAM"),
        ("Product Flow Rate", results.product_flow_kg_hr, "kg/hr", NUMBER_FORMAT, "BioSTEAM"),
        ("Feedstock Flow Rate", results.feedstock_flow_kg_hr, "kg/hr", NUMBER_FORMAT, "BioSTEAM"),
    ]
    for label, val, unit, fmt, src in metrics:
        _write_metric_row(ws, row, label, val, unit, fmt, src)
        row += 1

    row += 1
    _write_subheader(ws, row, "Capital Costs", 4)
    row += 1

    capex = [
        ("Installed Equipment Cost", results.installed_equipment_cost_usd / 1e6, "MM$"),
        ("Direct Permanent Investment (DPI)", results.dpi_usd / 1e6, "MM$"),
        ("Total Depreciable Capital (TDC)", results.tdc_usd / 1e6, "MM$"),
        ("Fixed Capital Investment (FCI)", results.fci_usd / 1e6, "MM$"),
        ("Total Capital Investment (TCI)", results.tci_usd / 1e6, "MM$"),
    ]
    for label, val, unit in capex:
        _write_metric_row(ws, row, label, val, unit, CURRENCY_M_FORMAT, "BioSTEAM")
        row += 1

    row += 1
    _write_subheader(ws, row, "Annual Operating Costs", 4)
    row += 1

    opex = [
        ("Material Costs", results.material_cost_usd_per_yr / 1e6, "MM$/yr"),
        ("Utility Costs", results.utility_cost_usd_per_yr / 1e6, "MM$/yr"),
        ("Fixed Operating Costs", results.foc_usd_per_yr / 1e6, "MM$/yr"),
        ("Total Annual Operating Cost", results.aoc_usd_per_yr / 1e6, "MM$/yr"),
    ]
    for label, val, unit in opex:
        _write_metric_row(ws, row, label, val, unit, CURRENCY_M_FORMAT, "BioSTEAM")
        row += 1

    row += 1
    _write_subheader(ws, row, "Operating Parameters", 4)
    row += 1
    _write_metric_row(ws, row, "Operating Days", results.operating_days, "days/yr", NUMBER_FORMAT)
    row += 1
    _write_metric_row(ws, row, "Operating Hours", results.operating_hours, "hr/yr", NUMBER_FORMAT)
    row += 1
    _write_metric_row(ws, row, "Plant Lifetime", results.plant_lifetime_years, "years")

    ws.freeze_panes = "A5"
    _auto_width(ws)


def _write_process_inputs_sheet(wb: Workbook, spec: ProcessSpec) -> None:
    """Sheet 2: Process Inputs — all parameters with source annotations."""
    ws = wb.create_sheet("Process Inputs")

    _write_header(ws, 1, ["Parameter", "Value", "Unit", "Source"])

    row = 2
    _write_subheader(ws, row, "Feedstock", 4)
    row += 1

    auto_filled = set(spec.metadata.auto_filled_params)

    feed = spec.feedstock
    params = [
        ("Feedstock Name", feed.name, "", "feedstock.name"),
        ("Flow Rate", feed.flow_rate_kg_hr, "kg/hr", "feedstock.flow_rate_kg_hr"),
        ("Price", feed.price_usd_per_ton, "$/ton", "feedstock.price_usd_per_ton"),
    ]
    for label, val, unit, path in params:
        source = "auto-filled" if path in auto_filled else "user"
        _write_metric_row(ws, row, label, val, unit, source=source)
        row += 1

    row += 1
    _write_subheader(ws, row, "Economic Parameters", 4)
    row += 1

    eco = spec.economic
    eco_params = [
        ("Operating Days", eco.operating_days, "days/yr", "economic.operating_days"),
        ("Plant Lifetime", eco.plant_lifetime_years, "years", "economic.plant_lifetime_years"),
        ("Discount Rate", eco.discount_rate, "", "economic.discount_rate"),
        ("Income Tax Rate", eco.income_tax_rate, "", "economic.income_tax_rate"),
        ("Depreciation", eco.depreciation_schedule, "", "economic.depreciation_schedule"),
        ("Construction Period", eco.construction_years, "years", "economic.construction_years"),
        ("Startup Period", eco.startup_months, "months", "economic.startup_months"),
        ("Working Capital Fraction", eco.working_capital_fraction, "of FCI", "economic.working_capital_fraction"),
    ]
    for label, val, unit, path in eco_params:
        source = "auto-filled" if path in auto_filled else "user"
        fmt = PERCENT_FORMAT if isinstance(val, float) and val < 1 else None
        _write_metric_row(ws, row, label, val, unit, fmt, source)
        row += 1

    row += 1
    _write_subheader(ws, row, "Unit Operation Parameters", 4)
    row += 1

    for unit_op in spec.units:
        ws.cell(row=row, column=1, value=f"{unit_op.id}: {unit_op.name or unit_op.type.value}").font = Font(bold=True)
        row += 1
        for key, val in unit_op.params.items():
            path = f"units.{unit_op.id}.params.{key}"
            source = "auto-filled" if path in auto_filled else "user"
            _write_metric_row(ws, row, f"  {key}", val, "", source=source)
            row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_mass_balance_sheet(wb: Workbook, spec: ProcessSpec) -> None:
    """Sheet 3: Mass Balance — stream connections (from ProcessSpec, not simulation)."""
    ws = wb.create_sheet("Mass Balance")

    _write_header(ws, 1, ["Stream", "From", "To", "Phase", "Components", "Flow Rate (kg/hr)"])

    for i, stream in enumerate(spec.streams, 2):
        ws.cell(row=i, column=1, value=f"S-{i-1:03d}")
        ws.cell(row=i, column=2, value=stream.from_id)
        ws.cell(row=i, column=3, value=stream.to_id)
        ws.cell(row=i, column=4, value=stream.phase or "-")
        ws.cell(row=i, column=5, value=", ".join(stream.components) if stream.components else "-")
        cell = ws.cell(row=i, column=6, value=stream.flow_rate_kg_hr)
        if stream.flow_rate_kg_hr:
            cell.number_format = NUMBER_FORMAT
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = THIN_BORDER

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_energy_balance_sheet(wb: Workbook, results: SimulationResults) -> None:
    """Sheet 4: Energy Balance — utility requirements per unit."""
    ws = wb.create_sheet("Energy Balance")

    _write_header(ws, 1, ["Unit ID", "Unit Type", "Heating (kW)", "Cooling (kW)", "Power (kW)"])

    total_heat = 0.0
    total_cool = 0.0
    total_power = 0.0

    for i, util in enumerate(results.utilities, 2):
        ws.cell(row=i, column=1, value=util.unit_id)
        ws.cell(row=i, column=2, value=util.unit_name)
        ws.cell(row=i, column=3, value=util.heating_duty_kW).number_format = NUMBER_FORMAT
        ws.cell(row=i, column=4, value=util.cooling_duty_kW).number_format = NUMBER_FORMAT
        ws.cell(row=i, column=5, value=util.power_kW).number_format = NUMBER_FORMAT
        total_heat += util.heating_duty_kW
        total_cool += util.cooling_duty_kW
        total_power += util.power_kW
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Totals row
    row = len(results.utilities) + 2
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_heat).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=4, value=total_cool).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=5, value=total_power).number_format = NUMBER_FORMAT
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_equipment_costs_sheet(wb: Workbook, results: SimulationResults) -> None:
    """Sheet 5: Equipment Costs — purchase + installed cost per unit."""
    ws = wb.create_sheet("Equipment Costs")

    _write_header(ws, 1, ["Unit ID", "Unit Type", "Purchase Cost ($)", "Installed Cost ($)", "Install Factor"])

    total_purchase = 0.0
    total_installed = 0.0

    sorted_costs = sorted(results.unit_costs, key=lambda x: x.installed_cost_usd, reverse=True)

    for i, uc in enumerate(sorted_costs, 2):
        ws.cell(row=i, column=1, value=uc.id)
        ws.cell(row=i, column=2, value=uc.unit_class)
        ws.cell(row=i, column=3, value=uc.purchase_cost_usd).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=uc.installed_cost_usd).number_format = CURRENCY_FORMAT
        factor = uc.installed_cost_usd / uc.purchase_cost_usd if uc.purchase_cost_usd > 0 else 0
        ws.cell(row=i, column=5, value=factor).number_format = "0.00"
        total_purchase += uc.purchase_cost_usd
        total_installed += uc.installed_cost_usd
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER

    row = len(sorted_costs) + 2
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_purchase).number_format = CURRENCY_FORMAT
    ws.cell(row=row, column=4, value=total_installed).number_format = CURRENCY_FORMAT
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_operating_costs_sheet(wb: Workbook, results: SimulationResults) -> None:
    """Sheet 6: Operating Costs — annual breakdown."""
    ws = wb.create_sheet("Operating Costs")

    _write_header(ws, 1, ["Cost Category", "Annual Cost (MM$/yr)", "% of Total"])

    total = abs(results.material_cost_usd_per_yr) + abs(results.utility_cost_usd_per_yr) + results.foc_usd_per_yr
    if total == 0:
        total = 1  # avoid division by zero

    costs = [
        ("Raw Materials", results.material_cost_usd_per_yr / 1e6),
        ("Utilities", results.utility_cost_usd_per_yr / 1e6),
        ("Fixed Operating Costs", results.foc_usd_per_yr / 1e6),
    ]

    for i, (label, val) in enumerate(costs, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val).number_format = CURRENCY_M_FORMAT
        ws.cell(row=i, column=3, value=abs(val * 1e6) / total).number_format = PERCENT_FORMAT
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = THIN_BORDER

    row = len(costs) + 2
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=results.aoc_usd_per_yr / 1e6).number_format = CURRENCY_M_FORMAT
    for col in range(1, 4):
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_cashflow_sheet(wb: Workbook, results: SimulationResults) -> None:
    """Sheet 7: Cash Flow — year-by-year from BioSTEAM cashflow table."""
    ws = wb.create_sheet("Cash Flow")

    cf = results.cashflow_table
    if cf is None:
        ws.cell(row=1, column=1, value="No cashflow data available")
        return

    # Write headers
    headers = ["Year"] + list(cf.columns)
    _write_header(ws, 1, headers)

    # Write data
    for i, (year, row_data) in enumerate(cf.iterrows(), 2):
        ws.cell(row=i, column=1, value=int(year) if hasattr(year, '__int__') else year)
        for j, val in enumerate(row_data, 2):
            cell = ws.cell(row=i, column=j, value=float(val) if hasattr(val, '__float__') else _safe_value(val))
            col_name = cf.columns[j - 2]
            if "MM$" in col_name or "cost" in col_name.lower():
                cell.number_format = CURRENCY_M_FORMAT
            elif "factor" in col_name.lower():
                cell.number_format = "0.0000"
            else:
                cell.number_format = NUMBER_FORMAT
            cell.border = THIN_BORDER

    ws.freeze_panes = "B2"
    _auto_width(ws)


def _write_sensitivity_sheet(wb: Workbook, spec: ProcessSpec) -> None:
    """Sheet 8: Sensitivity — placeholder for parameter sensitivity data.

    In Phase 2, this will contain Monte Carlo results and tornado diagram data.
    For MVP, it lists the key parameters that would be varied.
    """
    ws = wb.create_sheet("Sensitivity")

    ws.cell(row=1, column=1, value="Sensitivity Analysis Parameters").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value="(Full Monte Carlo analysis available in Phase 2)").font = SOURCE_FONT

    _write_header(ws, 4, ["Parameter", "Baseline Value", "Low (-20%)", "High (+20%)", "Impact on MESP"])

    # Key parameters for sensitivity
    sensitivity_params = [
        ("Feedstock Price", spec.feedstock.price_usd_per_ton or 83.0, "$/ton"),
        ("Operating Days", float(spec.economic.operating_days), "days/yr"),
        ("Enzyme Loading", 20.0, "mg/g cellulose"),
        ("Glucan-to-Glucose Conversion", 0.90, "fraction"),
        ("Xylose-to-Ethanol Conversion", 0.85, "fraction"),
        ("Glucose-to-Ethanol Conversion", 0.95, "fraction"),
        ("Discount Rate", spec.economic.discount_rate, "fraction"),
        ("Plant Lifetime", float(spec.economic.plant_lifetime_years), "years"),
        ("Construction Period", float(spec.economic.construction_years), "years"),
        ("Income Tax Rate", spec.economic.income_tax_rate, "fraction"),
    ]

    for i, (name, baseline, unit) in enumerate(sensitivity_params, 5):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=baseline)
        ws.cell(row=i, column=3, value=baseline * 0.8).number_format = NUMBER_FORMAT
        ws.cell(row=i, column=4, value=baseline * 1.2).number_format = NUMBER_FORMAT
        ws.cell(row=i, column=5, value="TBD — requires Monte Carlo").font = SOURCE_FONT
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER

    ws.freeze_panes = "A5"
    _auto_width(ws)


def write_tea_xlsx(
    results: SimulationResults,
    spec: ProcessSpec,
    path: str | Path,
) -> Path:
    """Generate the complete TEA XLSX report.

    Args:
        results: Simulation results from BioSTEAM
        spec: The ProcessSpec used for the simulation
        path: Output file path (.xlsx)

    Returns:
        Path to the saved XLSX file
    """
    path = Path(path)
    wb = Workbook()

    _write_summary_sheet(wb, results, spec)
    _write_process_inputs_sheet(wb, spec)
    _write_mass_balance_sheet(wb, spec)
    _write_energy_balance_sheet(wb, results)
    _write_equipment_costs_sheet(wb, results)
    _write_operating_costs_sheet(wb, results)
    _write_cashflow_sheet(wb, results)
    _write_sensitivity_sheet(wb, spec)

    wb.save(str(path))
    return path
